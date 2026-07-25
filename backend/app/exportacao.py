# backend/app/exportacao.py — Sistema Dono
#
# Módulo genérico de exportação tabular, usado por routes/relatorios.py
# para cumprir o `formato=pdf|xlsx` já documentado em api-endpoints.md §9
# desde a primeira versão da API, mas nunca implementado (só JSON existia
# até esta rodada).
#
# Decisão de design: PDF via reportlab (Platypus/Table) e XLSX via
# openpyxl, escrevendo VALORES já calculados (não fórmulas) — os
# relatórios daqui são fotografias de uma consulta que já rodou no
# Postgres (curva ABC, MRP, etc.), não planilhas para o usuário editar e
# recalcular; não há "modelo" para reabrir e mexer, então fórmula não
# agregaria nada e só complicaria (recalc, fórmulas pós-2007 sem suporte
# no LibreOffice etc.). Nenhuma biblioteca de template adicional — o
# texto que documentava esta lacuna já apontava "biblioteca de template,
# layout" como trabalho à parte; reportlab (Platypus) e openpyxl cobrem
# isso sozinhos para tabelas.
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Cor de cabeçalho única para PDF e XLSX — consistência visual entre os
# dois formatos exportados do mesmo relatório. Pública (sem "_") porque
# app/ficha_tecnica_pdf.py também a reaproveita, para a ficha técnica em
# PDF usar a mesma identidade visual dos relatórios tabulares.
COR_CABECALHO_HEX = "2C3E50"


def slugificar(texto: str) -> str:
    """'Curva ABC - Insumo/Gênero' -> 'curva_abc_insumo_genero'. Usado
    para montar o nome de arquivo do Content-Disposition — precisa ser
    ASCII-seguro (acentos e barras quebram alguns clientes HTTP em
    headers não codificados)."""
    texto = unicodedata.normalize("NFKD", texto.lower()).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto or "relatorio"


def _formatar_valor_texto(valor: Any) -> str:
    """Formata qualquer valor vindo de asyncpg/Python puro como texto
    legível para uma célula de PDF (Table do reportlab só aceita string/
    Paragraph, não os tipos nativos do Python)."""
    if valor is None:
        return "-"
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    if isinstance(valor, (Decimal, float)):
        texto = f"{float(valor):.4f}".rstrip("0").rstrip(".")
        return texto if texto not in ("", "-0") else "0"
    if isinstance(valor, (UUID,)):
        return str(valor)
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    if isinstance(valor, (list, tuple)):
        return ", ".join(str(v) for v in valor) if valor else "-"
    return str(valor)


def _valor_para_celula_xlsx(valor: Any):
    """openpyxl aceita nativamente str/int/float/bool/date/datetime —
    exceto que datetime CIENTE de fuso (tzinfo != None) faz o
    Workbook.save() levantar TypeError ("Excel does not support
    timezones in datetimes") — o formato XLSX não representa fuso
    horário em nenhuma célula de data. asyncpg devolve datetime com
    tzinfo para toda coluna TIMESTAMPTZ (ex.: classificacoes_abc.atualizado_em,
    usada em /relatorios/curva-abc) — bug real encontrado rodando contra
    Postgres de verdade (500 cru), não coberto pelos testes anteriores
    porque usavam datetime "ingênuo" (sem tzinfo). Descarta o tzinfo
    preservando a hora de parede (o valor já vem em UTC do banco — só
    perde a MARCAÇÃO de fuso na célula, não a hora em si).
    Também precisa de ajuda com Decimal (não suportado, vira float) e
    UUID/listas (viram string), o que Python-puro já não teria."""
    if valor is None:
        return None
    if isinstance(valor, datetime) and valor.tzinfo is not None:
        return valor.replace(tzinfo=None)
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, UUID):
        return str(valor)
    if isinstance(valor, (list, tuple)):
        return ", ".join(str(v) for v in valor) if valor else None
    return valor


def gerar_pdf_relatorio(titulo: str, secoes: list[dict], subtitulo: str | None = None) -> bytes:
    """secoes: lista de {"titulo": str, "colunas": [(rotulo, chave), ...],
    "linhas": [dict, ...]}. Um relatório com uma seção só (a maioria) só
    passa uma lista de 1 elemento; relatórios com mais de uma tabela
    (ex.: ruptura de estoque = lotes vencendo + insumos zerados) usam
    mais de uma seção no mesmo PDF, cada uma com seu próprio cabeçalho."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), title=titulo,
        topMargin=36, bottomMargin=36, leftMargin=28, rightMargin=28,
    )
    estilos = getSampleStyleSheet()
    historia = [Paragraph(titulo, estilos["Title"])]
    if subtitulo:
        historia.append(Paragraph(subtitulo, estilos["Normal"]))
    historia.append(Spacer(1, 12))

    for secao in secoes:
        if len(secoes) > 1:
            historia.append(Paragraph(secao["titulo"], estilos["Heading2"]))
            historia.append(Spacer(1, 6))

        colunas = secao["colunas"]
        linhas = secao["linhas"]
        cabecalho = [rotulo for rotulo, _ in colunas]
        dados = [cabecalho] + [
            [_formatar_valor_texto(linha.get(chave)) for _, chave in colunas] for linha in linhas
        ]

        tabela = Table(dados, repeatRows=1, hAlign="LEFT")
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{COR_CABECALHO_HEX}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        historia.append(tabela)

        if not linhas:
            historia.append(Spacer(1, 6))
            historia.append(Paragraph("Nenhum dado encontrado para os filtros aplicados.", estilos["Italic"]))

        historia.append(Spacer(1, 16))

    doc.build(historia)
    return buffer.getvalue()


def gerar_xlsx_relatorio(titulo: str, secoes: list[dict]) -> bytes:
    """Uma worksheet por seção — nome da aba vem de secao['titulo'],
    truncado em 31 caracteres (limite do formato XLSX) e sem os
    caracteres que o Excel proíbe em nome de aba."""
    wb = Workbook()
    wb.remove(wb.active)  # a aba padrão "Sheet" não deve sobrar vazia

    caracteres_proibidos = re.compile(r"[\[\]:*?/\\]")

    for secao in secoes:
        nome_aba = caracteres_proibidos.sub("_", secao["titulo"])[:31] or "Relatorio"
        ws = wb.create_sheet(title=nome_aba)

        colunas = secao["colunas"]
        linhas = secao["linhas"]
        ws.append([rotulo for rotulo, _ in colunas])
        for celula in ws[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill("solid", fgColor=COR_CABECALHO_HEX)

        for linha in linhas:
            ws.append([_valor_para_celula_xlsx(linha.get(chave)) for _, chave in colunas])

        # Largura automática — soma-se 2 de folga, com teto em 40 para
        # não deixar uma coluna de UUID esticando a planilha inteira.
        for col_cells in ws.columns:
            maior = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=0)
            ws.column_dimensions[col_cells[0].column_letter].width = min(maior + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
