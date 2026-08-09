# backend/tests/test_cotacao_import.py — Sistema Dono
#
# Suite de testes para o pipeline de importação de cotação por documento.
#
# Cobre:
#   A) cotacao_import.py (unit): extração de texto por formato (XML, XLSX),
#      parse JSON do LLM, matching de insumo/fornecedor, persistência em
#      cotacoes.
#   B) Endpoint POST /ia/importar-cotacao (integração): upload de documento,
#      enfileiramento assíncrono, validação de formato, tamanho, permissões.
#   C) Endpoint GET /ia/importar-cotacao/jobs/{job_id} (integração): polling
#      de status de job.
#   D) ai_worker.py branch COTACAO_DOCUMENTO: processamento assíncrono via
#      worker, tratamento de erro (arquivo ausente, ValueError no pipeline).
#
# Estratégia de mock:
#   - httpx.AsyncClient → mock do Ollama (LLM não deve rodar em CI)
#   - app.cotacao_import.get_pool → db_pool fixture (pool real de teste)
#   - app.routes.ia.enfileirar_job_cotacao_documento → mock no endpoint
#   - app.cotacao_import.processar_documento_cotacao → mock no worker
#     (import local no ai_worker → patch deve ser no módulo de origem)
#   - Redis → fixture mock_redis do conftest (já cobre set/get/close/delete)
#
# Fixtures do conftest utilizadas:
#   client, conn, db_pool, mock_redis
#   token_admin, token_chef, token_gestao, token_compras
#   usuario_admin (substitui usuario_sistema, que não existe neste projeto)
#
# PRÉ-REQUISITO: mesmos do conftest.py (banco dono_test com schema aplicado).

import io
import json
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from tests.conftest import auth_headers, err


# =====================================================================
# A) Testes unitários — extração de texto
# =====================================================================

class TestExtrairTextoXml:
    """Testa a extração de texto bruto de arquivos XML de cotação.

    Diferente do parser de NF-e (ElementTree por tag), cotações XML não
    têm schema fixo — o texto é passado diretamente ao LLM.
    """

    def test_xml_utf8(self):
        """XML em UTF-8 deve ser decodificado corretamente."""
        from app.cotacao_import import extrair_texto_xml
        xml = b"<cotacao><item>Frango</item></cotacao>"
        texto = extrair_texto_xml(xml)
        assert "Frango" in texto

    def test_xml_iso8859(self):
        """XML em ISO-8859-1 (comum em sistemas legados BR) deve ser
        decodificado corretamente como fallback."""
        from app.cotacao_import import extrair_texto_xml
        xml = "Frango Grill".encode("iso-8859-1")
        texto = extrair_texto_xml(xml)
        assert "Frango" in texto

    def test_xml_vazio_decodifica(self):
        """Bytes vazios devem resultar em string vazia sem exceção."""
        from app.cotacao_import import extrair_texto_xml
        xml = b""
        texto = extrair_texto_xml(xml)
        assert texto == ""


class TestExtrairTextoXlsx:
    """Testa a serialização de planilhas livres de fornecedor como texto tabular.

    Planilhas de cotação não seguem template fixo — o texto é serializado
    e enviado ao LLM para extração estruturada.
    """

    def _criar_xlsx(self, linhas: list) -> bytes:
        """Helper: cria um XLSX em memória com as linhas fornecidas."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in linhas:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_xlsx_simples(self):
        """Planilha com cabeçalho e dados deve ser serializada como texto
        com os valores de cada célula separados por pipe."""
        from app.cotacao_import import extrair_texto_xlsx
        xlsx = self._criar_xlsx([
            ["Produto", "Qtd", "Preco"],
            ["Arroz", 50, 3.50],
            ["Feijao", 30, 5.00],
        ])
        texto = extrair_texto_xlsx(xlsx)
        assert "Arroz" in texto
        assert "Feijao" in texto
        assert any(v in texto for v in ("3.5", "3,5", "3.50"))

    def test_xlsx_sem_dados_retorna_apenas_header_planilha(self):
        """Workbook sem células preenchidas não levanta exceção — retorna
        string com apenas o header da planilha. O openpyxl sempre gera
        ao menos uma linha vazia ao salvar, então o comportamento correto
        é não levantar ValueError (o LLM simplesmente não encontrará itens)."""
        from app.cotacao_import import extrair_texto_xlsx
        from openpyxl import Workbook
        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        texto = extrair_texto_xlsx(buf.getvalue())
        assert "Sheet" in texto or "Planilha" in texto

    def test_xlsx_respeita_max_linhas(self):
        """Planilha com muitas linhas deve ser truncada em max_linhas,
        indicando o truncamento no texto serializado."""
        from app.cotacao_import import extrair_texto_xlsx
        linhas = [["Item", str(i)] for i in range(500)]
        xlsx = self._criar_xlsx(linhas)
        texto = extrair_texto_xlsx(xlsx, max_linhas=10)
        assert "truncado" in texto


# =====================================================================
# A) Testes unitários — parse JSON do LLM
# =====================================================================

class TestParsearJsonLlm:
    """Testa o parser de resposta do LLM local (Ollama).

    O LLM pode retornar o JSON puro ou envolvido em markdown fences.
    O parser deve lidar com ambos e levantar ValueError para saídas inválidas.
    """

    def test_json_valido(self):
        """JSON puro retornado pelo LLM deve ser parseado corretamente."""
        from app.cotacao_import import _parsear_json_llm
        entrada = '[{"insumo_descricao": "Frango", "preco_unitario": 12.5}]'
        resultado = _parsear_json_llm(entrada)
        assert len(resultado) == 1
        assert resultado[0]["insumo_descricao"] == "Frango"

    def test_json_com_fences_markdown(self):
        """JSON envolto em ```json ... ``` deve ser extraído corretamente
        (alguns modelos Ollama sempre formatam assim)."""
        from app.cotacao_import import _parsear_json_llm
        entrada = '```json\n[{"insumo_descricao": "Sal", "preco_unitario": 1.0}]\n```'
        resultado = _parsear_json_llm(entrada)
        assert resultado[0]["insumo_descricao"] == "Sal"

    def test_json_invalido_levanta_valueerror(self):
        """Resposta não parseável deve levantar ValueError com mensagem
        que identifica o problema (para log do job)."""
        from app.cotacao_import import _parsear_json_llm
        with pytest.raises(ValueError, match="JSON"):
            _parsear_json_llm("isso nao e JSON")

    def test_json_nao_lista_levanta_valueerror(self):
        """LLM que retorna um objeto em vez de lista deve levantar ValueError.
        O pipeline espera sempre uma lista de itens."""
        from app.cotacao_import import _parsear_json_llm
        with pytest.raises(ValueError, match="lista"):
            _parsear_json_llm('{"insumo": "Frango"}')

    def test_lista_vazia_e_valida(self):
        """Lista vazia é JSON válido — a validação de nenhum item encontrado
        acontece mais acima, no orquestrador processar_documento_cotacao."""
        from app.cotacao_import import _parsear_json_llm
        resultado = _parsear_json_llm("[]")
        assert resultado == []


# =====================================================================
# A) Testes de integração — processar_documento_cotacao
# =====================================================================

@pytest.mark.asyncio
class TestProcessarDocumentoCotacao:
    """Testa o orquestrador principal do pipeline de importação.

    Usa o banco real (conn + db_pool fixtures) e mocka o LLM (httpx)
    e o pool global (app.cotacao_import.get_pool => db_pool).

    Princípio: nunca cria insumo ou fornecedor automaticamente. Itens
    sem associação clara retornam em itens_pendentes para revisão manual.
    """

    def _resposta_llm_mock(self, itens: list) -> AsyncMock:
        """Monta mock de httpx.AsyncClient que simula resposta do Ollama
        com os itens fornecidos serializados como JSON."""
        mock_response = MagicMock()
        mock_response.raise_for_status = MagicMock()
        mock_response.json.return_value = {"response": json.dumps(itens)}

        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.post = AsyncMock(return_value=mock_response)
        return mock_client

    async def test_item_associado_vira_cotacao(self, conn, db_pool, token_admin, client):
        """Item identificado pelo LLM que bate com insumo + fornecedor
        existentes deve ser persistido em cotacoes com origem=IA_IMPORTADA
        e status=PENDENTE_REVISAO."""
        cat_id = await conn.fetchval(
            "SELECT id FROM categorias WHERE nome = 'Secos e Despensa'"
        )
        ins_id = await conn.fetchval(
            """INSERT INTO insumos (nome, categoria_id, unidade, ativo)
               VALUES ('Arroz Cotacao Teste', $1, 'KG', TRUE) RETURNING id""",
            cat_id,
        )
        forn_id = await conn.fetchval(
            """INSERT INTO fornecedores (nome, ativo)
               VALUES ('Distribuidora Cotacao', TRUE) RETURNING id"""
        )

        itens_llm = [{
            "insumo_descricao": "Arroz Cotacao Teste",
            "quantidade": 50,
            "unidade": "KG",
            "preco_unitario": 4.50,
            "fornecedor_descricao": "Distribuidora Cotacao",
        }]

        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock(itens_llm)):
            resultado = await processar_documento_cotacao(
                formato="xml",
                arquivo_bytes=b"<cotacao><item>Arroz</item></cotacao>",
                usuario_id=uuid.uuid4(),
                fornecedor_hint_id=forn_id,
            )

        assert resultado["resumo"]["itens_criados"] == 1
        assert resultado["resumo"]["itens_pendentes"] == 0
        assert len(resultado["itens_criados"]) == 1
        assert resultado["itens_criados"][0]["preco_unitario"] == 4.50

        cotacao = await conn.fetchrow(
            "SELECT * FROM cotacoes WHERE insumo_id = $1", ins_id
        )
        assert cotacao is not None
        assert cotacao["origem"] == "IA_IMPORTADA"
        assert cotacao["status"] == "PENDENTE_REVISAO"
        assert float(cotacao["preco_unitario"]) == 4.50

    async def test_item_sem_insumo_vai_para_pendentes(self, client, conn, db_pool):
        """Item cujo nome nao bate com nenhum insumo cadastrado deve ir para
        itens_pendentes sem persistir nada em cotacoes."""
        forn_id = await conn.fetchval(
            """INSERT INTO fornecedores (nome, ativo)
               VALUES ('Fornecedor Pendente', TRUE) RETURNING id"""
        )

        itens_llm = [{
            "insumo_descricao": "Produto Inexistente XYZ 99999",
            "preco_unitario": 10.0,
            "fornecedor_descricao": "Fornecedor Pendente",
        }]

        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock(itens_llm)):
            resultado = await processar_documento_cotacao(
                formato="xml",
                arquivo_bytes=b"<cotacao/>",
                usuario_id=uuid.uuid4(),
                fornecedor_hint_id=forn_id,
            )

        assert resultado["resumo"]["itens_criados"] == 0
        assert resultado["resumo"]["itens_pendentes"] == 1
        assert "Insumo nao encontrado" in resultado["itens_pendentes"][0]["pendencia"] \
            or "Insumo não encontrado" in resultado["itens_pendentes"][0]["pendencia"]

    async def test_item_sem_fornecedor_vai_para_pendentes(self, client, conn, db_pool):
        """Item cujo fornecedor nao e identificavel pelo LLM e nao tem
        hint informado deve ir para itens_pendentes. O pipeline nao cria
        fornecedor automaticamente."""
        cat_id = await conn.fetchval(
            "SELECT id FROM categorias WHERE nome = 'Secos e Despensa'"
        )
        await conn.execute(
            """INSERT INTO insumos (nome, categoria_id, unidade, ativo)
               VALUES ('Feijao Sem Forn', $1, 'KG', TRUE)""",
            cat_id,
        )

        itens_llm = [{
            "insumo_descricao": "Feijao Sem Forn",
            "preco_unitario": 7.0,
            "fornecedor_descricao": None,
        }]

        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock(itens_llm)):
            resultado = await processar_documento_cotacao(
                formato="xml",
                arquivo_bytes=b"<cotacao/>",
                usuario_id=uuid.uuid4(),
                fornecedor_hint_id=None,
            )

        assert resultado["resumo"]["itens_pendentes"] >= 1
        pendencias = [i["pendencia"] for i in resultado["itens_pendentes"]]
        assert any("Fornecedor" in p for p in pendencias)

    async def test_item_sem_preco_e_ignorado(self, client, conn, db_pool):
        """Item sem preco_unitario identificado pelo LLM deve ser ignorado
        e ir para itens_pendentes."""
        itens_llm = [{
            "insumo_descricao": "Algo sem preco",
            "preco_unitario": None,
            "fornecedor_descricao": None,
        }]

        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock(itens_llm)):
            resultado = await processar_documento_cotacao(
                formato="xml",
                arquivo_bytes=b"<cotacao/>",
                usuario_id=uuid.uuid4(),
            )

        assert resultado["resumo"]["itens_criados"] == 0
        assert resultado["resumo"]["itens_pendentes"] == 1
        pendencia = resultado["itens_pendentes"][0]["pendencia"].lower()
        assert "pre" in pendencia  # "Preço" ou "preco"

    async def test_llm_retorna_lista_vazia_levanta(self, client, conn, db_pool):
        """LLM que nao identifica nenhum item deve levantar ValueError,
        que sera capturado pelo worker e gravado como erro no job."""
        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock([])):
            with pytest.raises(ValueError, match="Nenhum item"):
                await processar_documento_cotacao(
                    formato="xml",
                    arquivo_bytes=b"<cotacao/>",
                    usuario_id=uuid.uuid4(),
                )

    async def test_formato_invalido_levanta(self):
        """Formato nao suportado deve levantar ValueError antes mesmo de
        tentar extrair texto ou chamar o LLM."""
        from app.cotacao_import import processar_documento_cotacao

        with pytest.raises(ValueError, match="Formato nao suportado|Formato não suportado"):
            await processar_documento_cotacao(
                formato="docx",
                arquivo_bytes=b"qualquer coisa",
                usuario_id=uuid.uuid4(),
            )

    async def test_fornecedor_hint_bypassa_identificacao_llm(
        self, client, conn, db_pool
    ):
        """fornecedor_hint_id informado pelo usuario deve ser usado diretamente,
        ignorando o que o LLM identificou no texto do documento."""
        cat_id = await conn.fetchval(
            "SELECT id FROM categorias WHERE nome = 'Secos e Despensa'"
        )
        ins_id = await conn.fetchval(
            """INSERT INTO insumos (nome, categoria_id, unidade, ativo)
               VALUES ('Sal Hint Teste', $1, 'KG', TRUE) RETURNING id""",
            cat_id,
        )
        forn_id = await conn.fetchval(
            """INSERT INTO fornecedores (nome, ativo)
               VALUES ('Fornecedor Hint', TRUE) RETURNING id"""
        )

        itens_llm = [{
            "insumo_descricao": "Sal Hint Teste",
            "preco_unitario": 2.0,
            "fornecedor_descricao": "Outro Fornecedor Qualquer",
        }]

        from app.cotacao_import import processar_documento_cotacao

        with patch("app.cotacao_import.get_pool", return_value=db_pool), \
             patch("httpx.AsyncClient", return_value=self._resposta_llm_mock(itens_llm)):
            resultado = await processar_documento_cotacao(
                formato="xml",
                arquivo_bytes=b"<cotacao/>",
                usuario_id=uuid.uuid4(),
                fornecedor_hint_id=forn_id,
            )

        assert resultado["resumo"]["itens_criados"] == 1

        cotacao = await conn.fetchrow(
            "SELECT fornecedor_id FROM cotacoes WHERE insumo_id = $1", ins_id
        )
        assert cotacao["fornecedor_id"] == forn_id


# =====================================================================
# B) Testes de endpoint — POST /ia/importar-cotacao
# =====================================================================

@pytest.mark.asyncio
class TestImportarCotacaoEndpoint:
    """Testa o endpoint de upload de documento de cotacao.

    O endpoint enfileira um job assíncrono e retorna 202 com job_id.
    O processamento real (LLM + matching) acontece no ai_worker.
    """

    async def test_upload_pdf_retorna_202_e_job_id(self, client, token_admin):
        """Upload de PDF valido deve retornar 202 com job_id e status pendente."""
        pdf_bytes = b"%PDF-1.4 1 0 obj << /Type /Catalog >> endobj"
        files = {"arquivo": ("cotacao.pdf", io.BytesIO(pdf_bytes), "application/pdf")}

        with patch("app.routes.ia.enfileirar_job_cotacao_documento") as mock_enf:
            mock_enf.return_value = uuid.uuid4()
            r = await client.post(
                "/ia/importar-cotacao",
                headers=auth_headers(token_admin),
                files=files,
            )

        assert r.status_code == 202
        data = r.json()
        assert "job_id" in data
        assert data["status"] == "pendente"
        assert mock_enf.called

    async def test_upload_xml_retorna_202(self, client, token_admin):
        """Upload de XML valido deve retornar 202."""
        xml_bytes = b"<cotacao><item>Produto</item></cotacao>"
        files = {"arquivo": ("cotacao.xml", io.BytesIO(xml_bytes), "application/xml")}

        with patch("app.routes.ia.enfileirar_job_cotacao_documento") as mock_enf:
            mock_enf.return_value = uuid.uuid4()
            r = await client.post(
                "/ia/importar-cotacao",
                headers=auth_headers(token_admin),
                files=files,
            )

        assert r.status_code == 202
        assert mock_enf.called

    async def test_upload_xlsx_retorna_202(self, client, token_admin):
        """Upload de XLSX valido deve retornar 202."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Produto", "Preco"])
        ws.append(["Arroz", 4.50])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        files = {"arquivo": (
            "cotacao.xlsx", io.BytesIO(xlsx_bytes),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )}

        with patch("app.routes.ia.enfileirar_job_cotacao_documento") as mock_enf:
            mock_enf.return_value = uuid.uuid4()
            r = await client.post(
                "/ia/importar-cotacao",
                headers=auth_headers(token_admin),
                files=files,
            )

        assert r.status_code == 202
        assert mock_enf.called

    async def test_formato_invalido_retorna_400(self, client, token_admin):
        """Arquivo com extensao nao suportada deve retornar 400 com
        codigo FORMATO_INVALIDO."""
        files = {"arquivo": ("cotacao.docx", io.BytesIO(b"conteudo"), "application/octet-stream")}
        r = await client.post(
            "/ia/importar-cotacao",
            headers=auth_headers(token_admin),
            files=files,
        )
        assert r.status_code == 400
        assert err(r) == "FORMATO_INVALIDO"

    async def test_arquivo_grande_retorna_413(self, client, token_admin):
        """Arquivo acima de 10MB deve retornar 413 sem enfileirar job."""
        grande = b"x" * (10 * 1024 * 1024 + 1)
        files = {"arquivo": ("cotacao.pdf", io.BytesIO(grande), "application/pdf")}

        with patch("app.routes.ia.enfileirar_job_cotacao_documento"):
            r = await client.post(
                "/ia/importar-cotacao",
                headers=auth_headers(token_admin),
                files=files,
            )

        assert r.status_code == 413
        assert err(r) == "ARQUIVO_MUITO_GRANDE"

    async def test_sem_token_retorna_401(self, client):
        """Request sem token de autenticacao deve retornar 401."""
        files = {"arquivo": ("cotacao.pdf", io.BytesIO(b"%PDF"), "application/pdf")}
        r = await client.post("/ia/importar-cotacao", files=files)
        assert r.status_code == 401

    async def test_perfil_chef_negado(self, client, token_chef):
        """Perfil CHEF nao tem permissao para importar cotacoes
        (operacao restrita a COMPRAS e ADMIN)."""
        files = {"arquivo": ("cotacao.pdf", io.BytesIO(b"%PDF"), "application/pdf")}
        r = await client.post(
            "/ia/importar-cotacao",
            headers=auth_headers(token_chef),
            files=files,
        )
        assert r.status_code == 403

    async def test_fornecedor_id_invalido_retorna_400(self, client, token_admin):
        """fornecedor_id que nao e UUID valido deve retornar 400
        com codigo VALIDACAO_INVALIDA."""
        files = {"arquivo": ("cotacao.xml", io.BytesIO(b"<c/>"), "application/xml")}
        r = await client.post(
            "/ia/importar-cotacao?fornecedor_id=nao-e-uuid",
            headers=auth_headers(token_admin),
            files=files,
        )
        assert r.status_code == 400
        assert err(r) == "VALIDACAO_INVALIDA"

    async def test_fornecedor_id_valido_e_repassado(self, client, token_admin):
        """fornecedor_id UUID valido deve ser repassado como fornecedor_hint_id
        para o enfileirador."""
        forn_id = str(uuid.uuid4())
        files = {"arquivo": ("cotacao.xml", io.BytesIO(b"<c/>"), "application/xml")}

        with patch("app.routes.ia.enfileirar_job_cotacao_documento") as mock_enf:
            mock_enf.return_value = uuid.uuid4()
            r = await client.post(
                f"/ia/importar-cotacao?fornecedor_id={forn_id}",
                headers=auth_headers(token_admin),
                files=files,
            )

        assert r.status_code == 202
        hint = (
            mock_enf.call_args[0][3]
            if len(mock_enf.call_args[0]) > 3
            else mock_enf.call_args[1].get("fornecedor_hint_id")
        )
        assert str(hint) == forn_id


# =====================================================================
# C) Testes de endpoint — GET /ia/importar-cotacao/jobs/{job_id}
# =====================================================================

@pytest.mark.asyncio
class TestStatusJobCotacaoDocumento:
    """Testa o polling de status de jobs de importacao de cotacao.

    Segue o mesmo padrao de GET /ia/processar-nota/jobs/{job_id}:
    retorna status + resultado quando concluido, 404 para job inexistente,
    403 para job de outro usuario.
    """

    async def test_status_job_pendente(self, client, token_admin, conn, usuario_admin):
        """Job recem-criado deve retornar status pendente e o job_id correto."""
        job_id = await conn.fetchval(
            """INSERT INTO ia_jobs (tipo, solicitado_por, entrada)
               VALUES ('COTACAO_DOCUMENTO', $1, $2) RETURNING id""",
            usuario_admin["id"],
            json.dumps({"formato": "pdf"}),
        )
        r = await client.get(
            f"/ia/importar-cotacao/jobs/{job_id}",
            headers=auth_headers(token_admin),
        )
        assert r.status_code == 200
        data = r.json()
        assert data["status"] == "pendente"
        assert data["job_id"] == str(job_id)

    async def test_status_job_nao_encontrado(self, client, token_admin):
        """UUID inexistente deve retornar 404 com codigo RECURSO_NAO_ENCONTRADO."""
        r = await client.get(
            f"/ia/importar-cotacao/jobs/{uuid.uuid4()}",
            headers=auth_headers(token_admin),
        )
        assert r.status_code == 404
        assert err(r) == "RECURSO_NAO_ENCONTRADO"

    async def test_status_job_outro_usuario_retorna_403(
        self, client, token_admin, token_gestao, conn, usuario_admin
    ):
        """Job criado por um usuario nao pode ser consultado por outro usuario."""
        job_id = await conn.fetchval(
            """INSERT INTO ia_jobs (tipo, solicitado_por, entrada)
               VALUES ('COTACAO_DOCUMENTO', $1, $2) RETURNING id""",
            usuario_admin["id"],
            json.dumps({"formato": "xml"}),
        )
        r = await client.get(
            f"/ia/importar-cotacao/jobs/{job_id}",
            headers=auth_headers(token_gestao),
        )
        assert r.status_code == 403

    async def test_sem_token_retorna_401(self, client):
        """Polling de status sem autenticacao deve retornar 401."""
        r = await client.get(f"/ia/importar-cotacao/jobs/{uuid.uuid4()}")
        assert r.status_code == 401


# =====================================================================
# D) Testes do worker — branch COTACAO_DOCUMENTO
# =====================================================================

@pytest.mark.asyncio
class TestAiWorkerCotacaoDocumento:
    """Testa o branch COTACAO_DOCUMENTO no loop do ai_worker.

    O worker pega jobs pendentes do banco, recupera o arquivo do Redis
    e delega ao processar_documento_cotacao. Aqui mockamos o processamento
    real para testar apenas a orquestracao do worker.

    IMPORTANTE: processar_documento_cotacao e importado localmente dentro
    do ai_worker (import dentro do elif), entao o patch deve apontar para
    o modulo de origem (app.cotacao_import), nao para o modulo consumidor.
    Tentar patchar app.ai_worker.processar_documento_cotacao levanta
    AttributeError porque o nome nao existe como atributo do modulo.
    """

    async def test_job_cotacao_processado_com_sucesso(
        self, conn, db_pool, usuario_admin, mock_redis
    ):
        """Branch COTACAO_DOCUMENTO deve processar o job e marca-lo como
        concluido com o resultado serializado em ia_jobs.resultado."""
        from app.ai_worker import processar_proximo_job

        job_id = await conn.fetchval(
            """INSERT INTO ia_jobs (tipo, solicitado_por, entrada)
               VALUES ('COTACAO_DOCUMENTO', $1, $2) RETURNING id""",
            usuario_admin["id"],
            json.dumps({"formato": "xml", "fornecedor_id": None}),
        )

        mock_redis.get = AsyncMock(return_value=b"<cotacao/>")

        resultado_mock = {
            "status": "concluido",
            "formato": "xml",
            "itens_criados": [{"cotacao_id": str(uuid.uuid4()), "preco_unitario": 5.0}],
            "itens_pendentes": [],
            "resumo": {"total_itens": 1, "itens_criados": 1, "itens_pendentes": 0},
        }

        with patch(
            "app.cotacao_import.processar_documento_cotacao",
            new=AsyncMock(return_value=resultado_mock),
        ):
            processou = await processar_proximo_job(db_pool)

        assert processou is True

        job = await conn.fetchrow(
            "SELECT status, resultado FROM ia_jobs WHERE id = $1", job_id
        )
        assert job["status"] == "concluido"
        assert job["resultado"] is not None

    async def test_job_cotacao_sem_arquivo_redis_vai_para_erro(
        self, conn, db_pool, usuario_admin, mock_redis
    ):
        """Job COTACAO_DOCUMENTO sem arquivo no Redis (expirado ou nunca
        armazenado) deve ser marcado como erro com motivo descritivo."""
        from app.ai_worker import processar_proximo_job

        job_id = await conn.fetchval(
            """INSERT INTO ia_jobs (tipo, solicitado_por, entrada)
               VALUES ('COTACAO_DOCUMENTO', $1, $2) RETURNING id""",
            usuario_admin["id"],
            json.dumps({"formato": "pdf", "fornecedor_id": None}),
        )

        mock_redis.get = AsyncMock(return_value=None)

        processou = await processar_proximo_job(db_pool)
        assert processou is True

        job = await conn.fetchrow(
            "SELECT status, erro_motivo FROM ia_jobs WHERE id = $1", job_id
        )
        assert job["status"] == "erro"
        assert (
            "não encontrado" in job["erro_motivo"].lower()
            or "nao encontrado" in job["erro_motivo"].lower()
        )

    async def test_job_cotacao_erro_de_extracao_vai_para_erro(
        self, conn, db_pool, usuario_admin, mock_redis
    ):
        """ValueError no processamento deve marcar o job como erro com
        o motivo preservado, sem propagar a excecao para o loop do worker."""
        from app.ai_worker import processar_proximo_job

        job_id = await conn.fetchval(
            """INSERT INTO ia_jobs (tipo, solicitado_por, entrada)
               VALUES ('COTACAO_DOCUMENTO', $1, $2) RETURNING id""",
            usuario_admin["id"],
            json.dumps({"formato": "xml", "fornecedor_id": None}),
        )

        mock_redis.get = AsyncMock(return_value=b"<cotacao/>")

        with patch(
            "app.cotacao_import.processar_documento_cotacao",
            new=AsyncMock(
                side_effect=ValueError("Nenhum item de cotacao identificado")
            ),
        ):
            processou = await processar_proximo_job(db_pool)

        assert processou is True

        job = await conn.fetchrow(
            "SELECT status, erro_motivo FROM ia_jobs WHERE id = $1", job_id
        )
        assert job["status"] == "erro"
        assert "Nenhum item" in job["erro_motivo"]
