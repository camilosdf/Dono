# backend/app/cotacao_import.py — Sistema Dono
#
# Importação de cotação de fornecedor a partir de documentos (PDF, XML, XLSX).
# Diferente da NF-e (schema legal fixo), cotações não têm formato padronizado —
# a extração é assistida por LLM local (Ollama), não por parser determinístico.
#
# Fluxo:
#   1. Extrai texto bruto do documento (por formato).
#   2. Envia o texto ao LLM local pedindo extração estruturada em JSON.
#   3. Associa cada item a um insumo/fornecedor existentes via mesma
#      heurística de ILIKE já usada em app.nfe_xml (sem criar registros
#      automaticamente — fonte não é documento legal, exige revisão humana).
#   4. Itens associados viram `cotacoes` (origem='IA_IMPORTADA',
#      status='PENDENTE_REVISAO'). Itens não associados voltam como
#      itens_pendentes para associação manual, sem persistir nada.

import io
import os
import json
import uuid
import logging
from typing import Any, Dict, List, Optional

import httpx
import asyncpg

from app.database import get_pool
from app.nfe_xml import buscar_insumo_por_descricao, buscar_fornecedor_por_identificacao

logger = logging.getLogger(__name__)

OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://ollama:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.1:8b")

FORMATOS_SUPORTADOS = ("pdf", "xml", "xlsx")


# =====================================================================
# Extração de texto bruto por formato
# =====================================================================

def extrair_texto_xml(arquivo_bytes: bytes) -> str:
    """Decodifica XML como texto livre — cotações não seguem schema fixo
    como a NF-e, então não há parser por tag, apenas texto para o LLM."""
    for encoding in ("utf-8", "iso-8859-1", "cp1252"):
        try:
            return arquivo_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Não foi possível decodificar o XML com os encodings suportados")


def extrair_texto_xlsx(arquivo_bytes: bytes, max_linhas: int = 300) -> str:
    """Serializa uma planilha (livre, sem template fixo) como texto tabular
    simples, para servir de contexto ao LLM."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(arquivo_bytes), data_only=True, read_only=True)
    linhas_texto: List[str] = []

    for sheet in wb.worksheets:
        linhas_texto.append(f"--- Planilha: {sheet.title} ---")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_linhas:
                linhas_texto.append("... (truncado)")
                break
            valores = [str(v) if v is not None else "" for v in row]
            if any(v.strip() for v in valores):
                linhas_texto.append(" | ".join(valores))

    texto = "\n".join(linhas_texto)
    if not texto.strip():
        raise ValueError("Planilha vazia ou sem conteúdo legível")
    return texto


async def extrair_texto_documento(arquivo_bytes: bytes, formato: str) -> str:
    if formato == "pdf":
        from app.ocr import extrair_texto_pdf
        return await extrair_texto_pdf(arquivo_bytes)
    elif formato == "xml":
        return extrair_texto_xml(arquivo_bytes)
    elif formato == "xlsx":
        return extrair_texto_xlsx(arquivo_bytes)
    else:
        raise ValueError(
            f"Formato não suportado: {formato}. Use um de: {', '.join(FORMATOS_SUPORTADOS)}"
        )


# =====================================================================
# Extração estruturada via LLM local
# =====================================================================

def _parsear_json_llm(resposta_bruta: str) -> List[dict]:
    texto = resposta_bruta.strip()
    if texto.startswith("```"):
        texto = texto.split("```")[1]
        if texto.lower().startswith("json"):
            texto = texto[4:]
    texto = texto.strip()

    try:
        dados = json.loads(texto)
    except json.JSONDecodeError as e:
        raise ValueError(
            f"LLM não retornou JSON válido: {e}. Resposta bruta (truncada): {resposta_bruta[:500]}"
        )

    if not isinstance(dados, list):
        raise ValueError("LLM não retornou uma lista de itens")

    return dados


async def extrair_itens_cotacao_via_llm(texto: str, modelo: Optional[str] = None) -> List[dict]:
    """Envia o texto do documento ao LLM local pedindo extração estruturada
    dos itens de cotação em JSON. Trunca o texto para caber no contexto."""
    modelo = modelo or OLLAMA_MODEL
    texto_truncado = texto[:12000]

    prompt = f"""Você é um assistente que extrai dados de cotações de fornecedores de insumos alimentícios.

A seguir está o conteúdo bruto de um documento de cotação (PDF convertido em texto, XML ou planilha).
Extraia TODOS os itens de cotação encontrados e retorne APENAS um JSON válido, sem texto adicional,
sem explicações, sem marcação de código — apenas o array JSON puro.

Formato exato esperado (lista de objetos):
[
  {{
    "insumo_descricao": "nome/descrição do item conforme aparece no documento",
    "quantidade": <número ou null>,
    "unidade": "<unidade ou null>",
    "preco_unitario": <número decimal>,
    "fornecedor_descricao": "<nome do fornecedor, se identificável, ou null>"
  }}
]

Se não conseguir identificar o preço unitário de um item, NÃO o inclua na lista.

Documento:
{texto_truncado}

JSON:"""

    _timeout = float(os.getenv("OLLAMA_TIMEOUT", "120.0"))
    async with httpx.AsyncClient(timeout=_timeout) as client:
        response = await client.post(
            f"{OLLAMA_HOST}/api/generate",
            json={
                "model": modelo,
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.1, "top_p": 0.9},
            },
        )
        response.raise_for_status()
        resposta_bruta = response.json().get("response", "").strip()

    return _parsear_json_llm(resposta_bruta)


# =====================================================================
# Orquestração: extrai, associa, persiste
# =====================================================================

async def processar_documento_cotacao(
    formato: str,
    arquivo_bytes: bytes,
    usuario_id: uuid.UUID,
    fornecedor_hint_id: Optional[uuid.UUID] = None,
) -> Dict[str, Any]:
    """Processa um documento de cotação: extrai texto, chama o LLM para
    estruturar os itens, associa a insumos/fornecedores existentes e
    persiste os que puderem ser confiavelmente associados.

    Nunca cria insumo ou fornecedor automaticamente — a fonte (LLM sobre
    documento livre) não tem a confiabilidade legal de uma NF-e. Itens sem
    associação clara voltam em itens_pendentes para revisão manual.
    """
    if formato not in FORMATOS_SUPORTADOS:
        raise ValueError(
            f"Formato não suportado: {formato}. Use um de: {', '.join(FORMATOS_SUPORTADOS)}"
        )

    texto = await extrair_texto_documento(arquivo_bytes, formato)
    itens_extraidos = await extrair_itens_cotacao_via_llm(texto)

    if not itens_extraidos:
        raise ValueError("Nenhum item de cotação identificado no documento")

    pool = get_pool()
    itens_criados: List[dict] = []
    itens_pendentes: List[dict] = []

    async with pool.acquire() as conn:
        async with conn.transaction():
            fornecedor_id = fornecedor_hint_id

            if fornecedor_id is None:
                nomes_fornecedor = {
                    item.get("fornecedor_descricao")
                    for item in itens_extraidos
                    if item.get("fornecedor_descricao")
                }
                if len(nomes_fornecedor) == 1:
                    fornecedor_id = await buscar_fornecedor_por_identificacao(
                        conn, nome=next(iter(nomes_fornecedor)), criar_se_ausente=False,
                    )

            for item in itens_extraidos:
                preco = item.get("preco_unitario")
                descricao = item.get("insumo_descricao", "")

                if preco is None or not descricao:
                    itens_pendentes.append({
                        **item,
                        "pendencia": "Preço unitário ou descrição do item ausente",
                    })
                    continue

                insumo_id = await buscar_insumo_por_descricao(conn, descricao)
                if not insumo_id:
                    itens_pendentes.append({
                        **item,
                        "pendencia": "Insumo não encontrado — associação manual necessária",
                    })
                    continue

                if not fornecedor_id:
                    itens_pendentes.append({
                        **item,
                        "pendencia": "Fornecedor não identificado — associação manual necessária",
                    })
                    continue

                cotacao_id = await conn.fetchval(
                    """INSERT INTO cotacoes (insumo_id, fornecedor_id, preco_unitario, origem, status)
                       VALUES ($1, $2, $3, 'IA_IMPORTADA', 'PENDENTE_REVISAO')
                       RETURNING id""",
                    insumo_id, fornecedor_id, preco,
                )
                itens_criados.append({
                    "cotacao_id": str(cotacao_id),
                    "insumo_id": str(insumo_id),
                    "insumo_descricao": descricao,
                    "preco_unitario": preco,
                })

    return {
        "formato": formato,
        "itens_criados": itens_criados,
        "itens_pendentes": itens_pendentes,
        "resumo": {
            "total_itens": len(itens_extraidos),
            "itens_criados": len(itens_criados),
            "itens_pendentes": len(itens_pendentes),
        },
    }