# backend/tests/test_exportacao.py — Sistema Dono
#
# Cobre: geração de PDF e XLSX via API real (não teste isolado de
# exportacao.py). Valida Content-Type, assinatura binária do arquivo,
# e que datetime com tzinfo (bug real encontrado em produção — o
# asyncpg devolve TIMESTAMPTZ com tzinfo) não quebra o XLSX.
#
# ATUALIZAÇÃO (2026-07-24):
#   - Corrigido test_mrp_pdf_e_xlsx: adicionado parâmetro data_limite
#     obrigatório na chamada à rota /relatorios/mrp.
#   - Demais testes permanecem inalterados.

import io
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from tests.conftest import auth_headers, err


# ---------- Helpers de setup (reaproveitados de outras partes) ----------
async def _setup_dados(client, token_admin, token_chef, conn):
    """Cria insumo + lote + prato + refeição + confirma — mínimo para
    ter dados em classificacoes_abc (populadas pelo trigger na confirmação)."""
    cat_id = await conn.fetchval("SELECT id FROM categorias WHERE nome='Carnes, Aves e Peixes'")
    ins_r = await client.post("/insumos", headers=auth_headers(token_admin),
                              json={"nome": "Insumo Export", "categoria_id": str(cat_id), "unidade": "KG"})
    ins_id = ins_r.json()["id"]
    await client.post(f"/insumos/{ins_id}/lotes", headers=auth_headers(token_admin),
                      json={"valor_aquisicao": 60.0, "data_aquisicao": "2026-07-21", "quantidade": 20})

    prato_r = await client.post("/pratos", headers=auth_headers(token_chef),
                                json={"nome": "Prato Export", "genero_prato": "Prato Principal",
                                      "rendimento_base_porcoes": 4,
                                      "itens_receita": [{"insumo_id": ins_id, "tipo": "ALIMENTICIO",
                                                         "peso_bruto": 1, "fator_correcao": 1}]})
    prato_id = prato_r.json()["id"]

    ref_r = await client.post("/refeicoes", headers=auth_headers(token_chef),
                              json={"genero_refeicao": "Almoço Executivo", "data": "2026-08-10",
                                    "horario_inicio": "12:00", "horario_fim": "15:00", "qtd_pessoas": 4})
    ref_id = ref_r.json()["id"]
    await client.post(f"/refeicoes/{ref_id}/itens", headers=auth_headers(token_chef),
                      json={"prato_id": prato_id})
    await client.patch(f"/refeicoes/{ref_id}/confirmar", headers=auth_headers(token_chef))

    return ins_id, prato_id, ref_id


# ---------- Testes ----------
@pytest.mark.asyncio
class TestExportacaoRelatorios:

    async def test_curva_abc_json(self, client, token_admin, token_chef, conn):
        _, _, ref_id = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/relatorios/curva-abc?escopo=REFEICAO&id={ref_id}",
                             headers=auth_headers(token_admin))
        assert r.status_code == 200
        assert isinstance(r.json(), list)
        assert len(r.json()) > 0

    async def test_curva_abc_pdf(self, client, token_admin, token_chef, conn):
        _, _, ref_id = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/relatorios/curva-abc?escopo=REFEICAO&id={ref_id}&formato=pdf",
                             headers=auth_headers(token_admin))
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    async def test_curva_abc_xlsx_sem_erro_tzinfo(self, client, token_admin, token_chef, conn):
        """Regressão: TIMESTAMPTZ vindo do asyncpg (com tzinfo) não pode
        causar TypeError no openpyxl — bug real encontrado em produção."""
        _, _, ref_id = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/relatorios/curva-abc?escopo=REFEICAO&id={ref_id}&formato=xlsx",
                             headers=auth_headers(token_admin))
        assert r.status_code == 200, r.text
        ct = r.headers["content-type"]
        assert "spreadsheetml" in ct or "officedocument" in ct
        # Assinatura ZIP (PK) — todos os .xlsx são ZIP por dentro
        assert r.content[:2] == b"PK"

    async def test_mrp_pdf_e_xlsx(self, client, token_admin, token_chef, conn):
        await _setup_dados(client, token_admin, token_chef, conn)
        # CORREÇÃO: adicionar data_limite obrigatória
        data_limite = date.today() + timedelta(days=30)
        for fmt, magic in [("pdf", b"%PDF"), ("xlsx", b"PK")]:
            r = await client.get(
                f"/relatorios/mrp?formato={fmt}&data_limite={data_limite.isoformat()}",
                headers=auth_headers(token_admin)
            )
            assert r.status_code == 200, f"{fmt}: {r.text}"
            assert r.content[:2] == magic[:2]

    async def test_ruptura_estoque_xlsx_duas_abas(self, client, token_admin, token_chef, conn):
        """Ruptura de estoque gera 2 seções → 2 abas no XLSX.
        Nota: se este teste falhar com erro de tipo na query, verifique a
        implementação da rota em app/routes/relatorios.py (o parâmetro $1
        deve ser usado como integer, não string)."""
        await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get("/relatorios/ruptura-estoque?formato=xlsx",
                             headers=auth_headers(token_admin))
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert "Lotes Vencendo" in wb.sheetnames
        assert "Insumos Zerados" in wb.sheetnames

    async def test_abc_nao_calculado_retorna_404(self, client, token_admin):
        fake_id = "00000000-0000-0000-0000-000000000000"
        r = await client.get(f"/relatorios/curva-abc?escopo=REFEICAO&id={fake_id}",
                             headers=auth_headers(token_admin))
        assert r.status_code == 404
        assert err(r) == "ABC_NAO_CALCULADO"


@pytest.mark.asyncio
class TestExportacaoFichaTecnica:

    async def test_ficha_gerencial_json(self, client, token_admin, token_chef, conn):
        _, prato_id, _ = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/pratos/{prato_id}/ficha-tecnica?tipo=gerencial",
                             headers=auth_headers(token_chef))
        assert r.status_code == 200
        data = r.json()
        assert "custo_total_porcao" in data
        assert "ingredientes" in data

    async def test_ficha_gerencial_pdf(self, client, token_admin, token_chef, conn):
        _, prato_id, _ = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/pratos/{prato_id}/ficha-tecnica?tipo=gerencial&formato=pdf",
                             headers=auth_headers(token_chef))
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    async def test_ficha_insumo_pdf(self, client, token_admin, token_chef, conn):
        ins_id, prato_id, _ = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(
            f"/pratos/{prato_id}/ficha-tecnica?tipo=insumo&insumo_id={ins_id}&formato=pdf",
            headers=auth_headers(token_chef))
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    async def test_ficha_operacional_pdf_campos_nulos(self, client, token_admin, token_chef, conn):
        """Prato sem modo_preparo, instrucoes_apresentacao, equipamentos
        (campos opcionais todos nulos) não pode gerar 500."""
        _, prato_id, _ = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/pratos/{prato_id}/ficha-tecnica?tipo=operacional&formato=pdf",
                             headers=auth_headers(token_chef))
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    async def test_ficha_insumo_sem_insumo_id_retorna_400(self, client, token_admin, token_chef, conn):
        _, prato_id, _ = await _setup_dados(client, token_admin, token_chef, conn)
        r = await client.get(f"/pratos/{prato_id}/ficha-tecnica?tipo=insumo",
                             headers=auth_headers(token_chef))
        assert r.status_code == 400
        assert err(r) == "VALIDACAO_INVALIDA"